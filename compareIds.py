import openpyxl

# Scans both notebooks for duplicates and creates a copy of the second notebook with all of the duplicates removed
# workbookToKeep = str
# workbookToDeleteFrom = str
# outputName = str
# DO NOT include .xlsx at the end of any of any file names
def compareIds(workbookToKeep, workbookToDeleteFrom, outputName):
    '''
    A. Open both inputs and look for students sheet
    B. Loop through workbookToKeep and put them into a dictionary with vanId as the key. Value doesnt matter
    C. Loop through workbookToDeleteFrom and check for presence of vanId in dictionary
        1. If vanId exists
            a. clear this row
            b. increment duplicate count
    D. save file as outputName.xlsx
    '''
    wb1 = openpyxl.load_workbook(workbookToKeep + ".xlsx")
    wb2 = openpyxl.load_workbook(workbookToDeleteFrom + ".xlsx")
    sheet1 = wb1.get_sheet_by_name('students')
    sheet2 = wb2.get_sheet_by_name('students')
    start1 = 2
    end1  = sheet1.max_row + 1
    start2 = 2
    end2 = sheet2.max_row + 1

    names = {}

    for row in range(start1, end1):
        vanId = str(int(sheet1['A' + str(row)].value))
        # print (vanId)
        names[vanId] = row

    duplicates = 0

    for row in range(start2, end2):
        vanId = str(int(sheet2['A' + str(row)].value))
        if vanId in names:
            duplicates += 1
            sheet2['A' + str(row)].value = None
            sheet2['B' + str(row)].value = None
            sheet2['C' + str(row)].value = None
            sheet2['D' + str(row)].value = None
            sheet2['E' + str(row)].value = None

    wb2.save(outputName + ".xlsx")
    return duplicates

print ("Found " + str(compareIds('vcu', 'UoR', 'test')) + " duplicate names")
